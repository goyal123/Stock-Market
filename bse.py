# import openpyxl module
import openpyxl
from openpyxl.styles import Font, Color
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from xlwt import Workbook
from xlsxwriter import Workbook    
from nsetools import Nse
from bsedata.bse import BSE
import smtplib
import pandas as pd
from nsepy import get_history
from datetime import date
import matplotlib.pyplot as plt


def bsemarket():
    path = "C:\\Users\\Akhil\\Desktop\\Stock Market\\Stock-Book\\Stock-Book.xlsx"
    bse = BSE()
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    max_row = sheet_obj.max_row    
    max_col = sheet_obj.max_column
    
    message_bse =''
    for i in range(1, max_row-1):
        company =  str(sheet_obj.cell(row = i+1, column = 1).value)
        exchange = str(sheet_obj.cell(row = i+1, column = 2).value)
        quote_name = str(sheet_obj.cell(row = i+1, column = 3).value)
        cut_off_price = sheet_obj.cell(row = i+1, column = 4).value

        if exchange=='BSE':
            quote = bse.getQuote(quote_name)
            sheet_obj.cell(row = i+1, column = 5).value=int(float(quote['currentValue']))
            if ((sheet_obj.cell(row = i+1, column = 5).value)-(sheet_obj.cell(row = i+1, column = 4).value))>=15:
                sheet_obj.cell(row = i+1, column = 6).value='BUY'
                sheet_obj.cell(row = i+1, column = 6).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type = "solid")
                message_bse=message_bse+company+" - "+str(quote['currentValue'])+" | "
        
    wb_obj.save('Stock-book.xlsx')
    return message_bse    
