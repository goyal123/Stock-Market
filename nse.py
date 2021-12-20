# import openpyxl module
import openpyxl
from openpyxl.styles import Font, Color
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from xlwt import Workbook
from xlsxwriter import Workbook    
from nsetools import Nse
from bsedata.bse import BSE
import smtplib
import pandas as pd
from nsepy import get_history
from datetime import date
import matplotlib.pyplot as plt


def nsemarket():
    path = "C:\\Users\\Akhil\\Desktop\\Stock Market\\Stock-Book\\Stock-Book.xlsx"
    nse = Nse()
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active

    max_row = sheet_obj.max_row    
    max_col = sheet_obj.max_column
    message='STOCKS TO BUY:\n'

    for i in range(1, max_row-1):
        company =  str(sheet_obj.cell(row = i+1, column = 1).value)
        exchange = str(sheet_obj.cell(row = i+1, column = 2).value)
        quote_name = str(sheet_obj.cell(row = i+1, column = 3).value)
        cut_off_price = sheet_obj.cell(row = i+1, column = 4).value

        if exchange=='NSE':
            quote = nse.get_quote(quote_name)
            sheet_obj.cell(row = i+1, column = 5).value=int(quote['lastPrice'])
            if ((sheet_obj.cell(row = i+1, column = 5).value)-(sheet_obj.cell(row = i+1, column = 4).value))>=15:
                sheet_obj.cell(row = i+1, column = 6).value='BUY'
                sheet_obj.cell(row = i+1, column = 4).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
                sheet_obj.cell(row = i+1, column = 5).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
                sheet_obj.cell(row = i+1, column = 6).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type = "solid")
                message=message+company+" - "+str(quote['lastPrice'])+" | "
            else:
                sheet_obj.cell(row = i+1, column = 6).value=''
                sheet_obj.cell(row = i+1, column = 1).fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid")
                sheet_obj.cell(row = i+1, column = 4).fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid")
                sheet_obj.cell(row = i+1, column = 5).fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid")
                sheet_obj.cell(row = i+1, column = 6).fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type = "solid")
        
    wb_obj.save('Stock-Book.xlsx')
    return message
        
